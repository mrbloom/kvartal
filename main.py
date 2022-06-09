import csv
import datetime
import re
from glob import glob
from openpyxl import load_workbook

def extract_date(filename):
    return re.search(r"\d{4}\.\d{2}.\d{2}",filename).group(0)

def split_array(arr,delimiter=(None, None, None, None, None, None, None, None, None, None, None)):
    get_indexes = lambda x, xs: [i for (y, i) in zip(xs, range(len(xs))) if x == y]
    indexes = get_indexes(delimiter,arr)
    indexes.append(len(arr))
    indexes.insert(0,-1)
    # block_indexes = [ [indexes[i]+1,indexes[i+1]] for i in range(len(indexes)-1) ]
    blocks =  [ arr[indexes[i]+1:indexes[i+1]] for i in range(len(indexes)-1) ]
    return blocks

#2022.06.01 12:14:00;2022.06.01 12:17:00;1;00:00:15;*;111
def reduce_row(row, block_id:int):
    [hour,minute,second] = row[3].hour, row[3].minute, row[3].second
    prev_dt = row[0] + datetime.timedelta(hours=hour,minutes=minute,seconds=second) - datetime.timedelta(minutes=15)
    prev_date_time = prev_dt.strftime("%Y.%m.%d %H:%M:%S")
    day_date = str(row[0].strftime("%Y.%m.%d"))
    block_time = str(row[3].strftime("%H:%M:%S"))
    id = row[7]
    duration = str ( datetime.timedelta(seconds=int(row[8])) )
    return [ f"{prev_date_time}", f"{day_date} {block_time}", f"{id}.ts", duration, "*", str(block_id) ]


def convert_xlsx(filename):
    wb = load_workbook(filename)
    sheet = wb.sheetnames[0]
    wb_values = [value for value in wb[sheet].values]
    blocks = split_array(wb_values)

    filedate = extract_date(filename)

    csvfilename = "Kvartal-" + filedate + ".csv"
    with open(csvfilename, 'w', newline='') as csvfile:
       writer = csv.writer(csvfile, delimiter=';')

       intdate = int(filedate.replace(".",""))

       for (i,block) in enumerate(blocks):
           blockid = intdate * 1000 + i + 1

           for row in block:
               writer.writerow(reduce_row(row,blockid))

for xlsx_filename in glob("*.xlsx"):
    convert_xlsx(xlsx_filename)


